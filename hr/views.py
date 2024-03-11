import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from users.models import *
from vacancies.models import *
from django.contrib import messages
from .forms import *
from django.db.models import Q
from excel_response import ExcelResponse
from django.http import HttpResponse
import openpyxl
from django.db.models import Count
import json
import pyperclip
from django.db.models import Q, Count
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.urls import reverse
from django.contrib.sites.shortcuts import get_current_site
import openpyxl
from users.models import Staff
from users.forms import *
from django.core.paginator import *
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponseForbidden
from users.decorators import *
from openpyxl.styles import Alignment


@admins
def dashboard(request):
    return render(request, 'hr/dashboard.html')


@admins
def system_users(request):
    users = CustomUser.objects.filter(access_level=0)
    return render(request, 'hr/users.html', {'users': users})


@admins
def job_types(request):
    job_types_list = JobType.objects.all()
    return render(request, 'hr/job_types.html', {'job_types': job_types_list})


@admins
def create_job_type(request):
    if request.method == 'POST':
        form = JobTypeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('hr:job_types')
    else:
        form = JobTypeForm()

    return render(request, 'hr/create_job_type.html', {'form': form})


def edit_job_type(request, job_type_id):
    job_type = get_object_or_404(JobType, id=job_type_id)

    if request.method == 'POST':
        form = JobTypeForm(request.POST, request.FILES, instance=job_type)
        if form.is_valid():
            form.save()
            return redirect('hr:job_types')
    else:
        form = JobTypeForm(instance=job_type)

    return render(request, 'hr/edit_job_type.html', {'form': form, 'job_type': job_type})


def delete_job_type(request, job_type_id):
    job_type = get_object_or_404(JobType, id=job_type_id)

    job_type.delete()
    return redirect('hr:job_types')


@admins
def jobs(request):
    search_query = request.GET.get('search')
    job_discipline_filter = request.GET.get('job_discipline')
    vacancy_type_filter = request.GET.get('vacancy_type')

    jobs = Vacancy.objects.all()

    if search_query:
        jobs = jobs.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if job_discipline_filter:
        # Ensure you filter by the correct field name ('job_discipline_id')
        jobs = jobs.filter(job_discipline_id=job_discipline_filter)

    if vacancy_type_filter:
        # Ensure you filter by the correct field name ('job_type_id')
        jobs = jobs.filter(job_type_id=vacancy_type_filter)

    job_disciplines = JobDiscipline.objects.all()

    # Include JobTypes for 'vacancies' context variable
    vacancy_types = JobType.objects.all()

    context = {
        'jobs': jobs,
        'search_query': search_query,
        'selected_job_discipline': job_discipline_filter,
        'selected_vacancy_type': vacancy_type_filter,
        'job_disciplines': job_disciplines,
        'vacancy_types': vacancy_types,
    }

    return render(request, 'hr/jobs.html', context)


@system_admin_hr_post_required
def edit_job(request, job_id):
    job = get_object_or_404(Vacancy, pk=job_id)

    if request.method == 'POST':
        form = VacancyForm(request.POST, instance=job)
        if form.is_valid():
            user = request.user
            job.last_updated_by = f'{user.first_name} {user.last_name}' if user.first_name and user.last_name else user.username
            form.save()
            messages.success(request, 'Vacancy updated successfully')
            return redirect('hr:jobs')
        else:
            # Add error messages for invalid form data
            messages.error(
                request, 'Error updating vacancy. Please correct the errors below.')
    else:
        form = VacancyForm(instance=job)

    return render(request, 'hr/edit_job.html', {'form': form, 'job': job})


@system_admin_hr_post_required
def delete_job(request, job_id):
    job = get_object_or_404(Vacancy, pk=job_id)
    job.delete()
    messages.success(request, 'Vacancy Deleted successfully')
    return redirect('hr:jobs')


@system_admin_hr_post_required
def create_job(request):
    if request.method == 'POST':
        form = VacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save(commit=False)
            user = request.user
            vacancy.created_by = f'{user.first_name} {user.last_name}' if user.first_name and user.last_name else user.username
            vacancy.save()
            messages.success(
                request, f'Vacancy "{vacancy.title}" created successfully.')
            return redirect('hr:jobs')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error in {field}: {error}')
    else:
        form = VacancyForm()

    context = {'form': form}
    return render(request, 'hr/create_job.html', context)


@admins
def job_detail(request, vacancy_id):
    job = get_object_or_404(Vacancy, pk=vacancy_id)

    context = {'job': job}
    return render(request, 'hr/job_detail.html', context)


# @system_admin_hr_publish_required
@system_admin_hr_post_required
def publish(request, job_id):
    job = Vacancy.objects.get(pk=job_id)
    job.published = not job.published
    job.save()
    return redirect('hr:jobs')


@admins
def applications(request):
    search_query = request.GET.get('search')
    job_discipline_filter = request.GET.get('job_discipline')
    vacancy_type_filter = request.GET.get('vacancy_type')
    # Retrieve all vacancies
    jobs = Vacancy.objects.all()

    if search_query:
        jobs = jobs.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if job_discipline_filter:
        # Ensure you filter by the correct field name ('job_discipline_id')
        jobs = jobs.filter(job_discipline_id=job_discipline_filter)

    if vacancy_type_filter:
        # Ensure you filter by the correct field name ('vacancy_type')
        jobs = jobs.filter(vacancy_type=vacancy_type_filter)

    # Retrieve all applications
    applications = Application.objects.all()

    job_disciplines = JobDiscipline.objects.all()

    context = {
        'jobs': jobs,
        'applications': applications,
        'selected_vacancy_type': vacancy_type_filter,
        'search_query': search_query,
        'job_disciplines': job_disciplines,
        'vacancy_types': Vacancy.VACANCY_TYPES,
    }

    return render(request, 'hr/applications.html', context)


@admins
def application_detail(request, vacancy_id, filter_criteria=None):
    # Retrieve the vacancy object or return a 404 if it doesn't exist
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)

    # Retrieve all applications related to this vacancy
    applications = Application.objects.filter(vacancy=vacancy)

    educational_levels = EducationalLevel.objects.all()

    # Apply filters based on the filter criteria if it's provided
    if filter_criteria:
        if filter_criteria == 'qualified':
            applications = applications.filter(qualify=True)
        elif filter_criteria == 'disqualified':
            applications = applications.filter(qualify=False)
        elif filter_criteria == 'shortlisted':
            applications = applications.filter(shortlisted=True)

    # Additional filters
    education_level = request.GET.get('education_level')
    ethnicity = request.GET.get('ethnicity')
    gender = request.GET.get('gender')
    disability = request.GET.get('disability')
    experience = request.GET.get('experience')
    name_search = request.GET.get('name_search')

    # Apply additional filters
    if education_level:
        applications = applications.filter(
            applicant__resume__educational_level=education_level)
    if ethnicity:
        applications = applications.filter(
            applicant__resume__ethnicity=ethnicity)
    if gender:
        applications = applications.filter(applicant__resume__gender=gender)
    if disability:
        applications = applications.filter(
            applicant__resume__disability=disability)
    if experience:
        applications = applications.filter(
            applicant__work_experiences__years__gte=experience)
    if name_search:
        applications = applications.filter(
            Q(applicant__first_name__icontains=name_search) |
            Q(applicant__last_name__icontains=name_search)
        )

    export_excel = request.GET.get('export_excel')
    if export_excel:
        # Create a list of dictionaries representing the data you want to export
        data = []
        for application in applications:
            user = application.applicant
            try:
                resume = user.resume
            except Resume.DoesNotExist:
                # Create a new Resume if it doesn't exist
                # Combine first and last name
                full_name = f"{user.first_name} {user.last_name}"
                resume = Resume.objects.create(
                    user=user,
                    full_name=full_name,
                    email_address=user.email,
                )

            basic_education = user.basic_education.first()
            further_studies = user.further_studies.first()
            work_experience = user.work_experiences.all()[:3]
            certifications = user.certifications.all()[:3]
            memberships = user.memberships.all()[:3]
            referees = user.referees.all()[:3]

            print(referees)

            full_name = resume.full_name
            username = user.username  # Add this line to get the username

            application_data = {
                # Include username in the "Name" field
                'Username/Staff No.': username,
                'Full Name': full_name,
                'Gender': resume.gender if resume.gender else '',
                'Disability': 'Yes' if resume.disability else 'No',
                'Ethnicity': resume.ethnicity.name if resume.ethnicity else '',
                'Highest Educational Level': resume.educational_level.name if resume.educational_level else '',
                'High School': '',
                'College/University': '',
                'Professional Certifications': '',
                'Professional Membership': '',
                'Work Experience': '',
                'Referees': '',
                'Date Applied': application.application_date.strftime("%Y-%m-%d %H:%M:%S"),
                'Reference Number': application.reference_number,
                'Status': 'Qualified' if application.qualify else 'Not Qualified',
                'Shortlisted': 'Yes' if application.shortlisted else 'No',
            }

            if basic_education:
                # Handle Academic Details data
                academic_data = ''
                institution_name = f"School: {basic_education.name_of_the_school}"
                admission_number = f"Index_number: {basic_education.index_number}"
                start_year = f"Start Year: {basic_education.date_started}"
                end_year = f"End Year: {basic_education.date_ended}"
                grade = f"Grade attained: {basic_education.grade_attained}"
                academic_data += f"{institution_name}\n{admission_number}\n{start_year}\n{end_year}\n{grade}\n\n"

                application_data['High School'] = academic_data

            if further_studies:
                # Handle Further Studies data
                further_studies_data = f"Institution Name: {further_studies.institution_name}\n" \
                    f"Certification: {further_studies.certifications.certification_name if further_studies.certifications else ''}\n" \
                    f"Course Undertaken: {further_studies.course_undertaken}\n" \
                    f"Start Date: {further_studies.date_started.strftime('%Y-%m-%d') if further_studies.date_started else ''}\n" \
                    f"End Date: {further_studies.date_ended.strftime('%Y-%m-%d') if further_studies.date_ended else ''}\n" \
                    f"Grade: {further_studies.grade}"

                application_data['College/University'] = further_studies_data

            if certifications:
                # Handle Certifications data
                certifications_data = ''
                for cert in certifications:
                    name = f"Name: {cert.name}"
                    certifying_body_name = f"Certifying Body: {cert.certifying_body.name if cert.certifying_body else ''}"
                    date_attained = f"Date Attained: {cert.date_attained.strftime('%Y-%m-%d') if cert.date_attained else ''}"

                    certifications_data += f"{name}\n{certifying_body_name}\n{date_attained}\n\n"

                application_data['Professional Certifications'] = certifications_data

            if certifications:
                # Handle Certifications data
                certifications_data = ''
                for cert in certifications:
                    name = f"Name: {cert.name}"
                    certifying_body_name = f"Certifying Body: {cert.certifying_body.name if cert.certifying_body else ''}"
                    date_attained = f"Date Attained: {cert.date_attained.strftime('%Y-%m-%d') if cert.date_attained else ''}"

                    certifications_data += f"{name}\n{certifying_body_name}\n{date_attained}\n\n"

                application_data['Professional Certifications'] = certifications_data

            if memberships:
                # Handle Memberships data
                memberships_data = ''
                for membership in memberships:
                    title = f"Membership Title: {membership.membership_title}"
                    number = f"Membership Number: {membership.membership_number}"
                    body = f"Membership Body: {membership.membership_body}"
                    joined_date = f"Date Joined: {membership.date_joined.strftime('%Y-%m-%d') if membership.date_joined else ''}"

                    memberships_data += f"{title}\n{number}\n{body}\n{joined_date}\n\n"

                application_data['Professional Membership'] = memberships_data

            if work_experience:
                # Handle Work Experience data
                work_experience_data = ''

                total_years = 0
                total_months = 0

                for experience in work_experience:
                    company_name = f"Company Name: {experience.company_name}"
                    position = f"Position: {experience.position}"
                    start_date = experience.date_started.strftime(
                        '%Y-%m-%d') if experience.date_started else ''
                    end_date = experience.date_ended.strftime(
                        '%Y-%m-%d') if experience.date_ended else 'In Progress' if experience.currently_working else ''
                    company_address = f"Company Address: {experience.company_address}"
                    company_phone = f"Company Phone: {experience.company_phone}"
                    responsibilities = f"Responsibilities: {experience.responsibilities}"

                    work_experience_data += f"{company_name}\n{position}\nStart Date: {start_date}\nEnd Date: {end_date}\n{company_address}\n{company_phone}\n{responsibilities}\n"

                    # Calculate the duration of each experience
                    if experience.date_started and experience.date_ended:
                        delta = experience.date_ended - experience.date_started
                        years_worked = delta.days // 365
                        months_worked = (delta.days % 365) // 30

                        work_experience_data += f"Years Worked: {years_worked} years\nMonths Worked: {months_worked} months\n\n"

                        # Accumulate for total
                        total_years += years_worked
                        total_months += months_worked

                    # Add space between instances
                    work_experience_data += "\n"

                total_experience = f"Total Experience: {total_years} years and {total_months} months"
                work_experience_data += total_experience

                application_data['Work Experience'] = work_experience_data

            if referees:
                # Handle Referees data
                referees_data = ''
                for referee in referees:
                    full_name = f"Full Name: {referee.full_name}"
                    organization = f"Organization: {referee.organization}"
                    designation = f"Designation: {referee.designation}"
                    phone = f"Phone: {referee.phone}"
                    email = f"Email: {referee.email}"

                    referees_data += f"{full_name}\n{organization}\n{designation}\n{phone}\n{email}\n\n"

                application_data['Referees'] = referees_data

            data.append(application_data)

        # Define column headers for Excel export
        headers = [
            'Username/Staff No.', 'Full Name', 'Gender', 'Disability', 'Ethnicity',
            'Highest Educational Level',
            'High School',
            'College/University',
            'Professional Certifications',
            'Professional Membership',
            'Work Experience',
            'Referees',
            'Date Applied', 'Reference Number', 'Status', 'Shortlisted',
        ]

        # Create a workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

       # Write title
        title = f"{application.vacancy.title} / {application.vacancy.ref} Applications"
        title_row = ws.cell(row=1, column=8, value=title)
        title_row.alignment = Alignment(horizontal='center')
        title_row.font = openpyxl.styles.Font(size=14, bold=True)

        # Write headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_num, value=header)
            # Adjust the width of columns as needed
            if header == 'Highest Educational Level':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20
            if header == 'High School':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20
            if header == 'Full Name':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20
            if header == 'College/University':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20
            if header == 'Professional Certifications':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20

            if header == 'Professional Membership':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20

            if header == 'Work Experience':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20

            if header == 'Referees':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 20

        # Write data
        for row_num, application_data in enumerate(data, 3):
            for col_num, value in enumerate(application_data.values(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Adjust alignment and format for specific columns if needed
                if headers[col_num - 1] == 'High School':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'College/University':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Professional Certifications':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Professional Membership':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Work Experience':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Referees':
                    cell.alignment = Alignment(wrap_text=True)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="applications for {application.vacancy.title}-{application.vacancy.ref}.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    context = {
        'vacancy': vacancy,
        'educational_levels': educational_levels,
        'applications': applications,
    }

    return render(request, 'hr/application_detail.html', context)


@system_admin_required
def toggle_shortlist(request, vacancy_id, application_id):
    application = get_object_or_404(Application, pk=application_id)

    # Toggle the shortlisted status
    application.shortlisted = not application.shortlisted
    application.save()

    # Redirect to the application detail page for the specified vacancy_id
    return redirect('hr:application_detail', vacancy_id=vacancy_id)


@system_admin_hr_required
def resume(request, user_id):
    # Retrieve the user (applicant) and related information
    applicant = get_object_or_404(CustomUser, pk=user_id)

    try:
        resume = Resume.objects.get(user=applicant)
    except Resume.DoesNotExist:
        resume = None  # Set resume to None if it doesn't exist

    try:
        basic_education = BasicEducation.objects.get(user=applicant)
    except BasicEducation.DoesNotExist:
        basic_education = None  # Set basic_education to None if it doesn't exist

    try:
        further_studies = FurtherStudies.objects.get(user=applicant)
    except FurtherStudies.DoesNotExist:
        further_studies = None  # Set further_studies to None if it doesn't exist

    try:
        memberships = Membership.objects.filter(user=applicant)
    except Membership.DoesNotExist:
        memberships = []  # Set memberships to an empty list if they don't exist

    try:
        work_experiences = WorkExperience.objects.filter(user=applicant)
    except WorkExperience.DoesNotExist:
        work_experiences = []  # Set work_experiences to an empty list if they don't exist

    try:
        referees = Referee.objects.filter(user=applicant)
    except Referee.DoesNotExist:
        referees = []  # Set referees to an empty list if they don't exist

    try:
        certifications = Certification.objects.filter(user=applicant)
    except Certification.DoesNotExist:
        certifications = []  # Set certifications to an empty list if they don't exist

    try:
        objective = ProfessionalSummary.objects.get(user=applicant)
    except ProfessionalSummary.DoesNotExist:
        objective = None  # Set objective to None if it doesn't exist

    context = {
        'applicant': applicant,
        'resume': resume,
        'basic_education': basic_education,
        'further_studies': further_studies,
        'memberships': memberships,
        'work_experience': work_experiences,
        'Referees': referees,
        'certifications': certifications,
        'objective': objective,
    }

    return render(request, 'hr/resume.html', context)


@admins
def create_job_discipline(request):
    if request.method == 'POST':
        form = JobDisciplineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Job discipline added succesfully")
            # Redirect to a success page after creating the discipline
            return redirect('hr:job_disciplines')
    else:
        form = JobDisciplineForm()

    return render(request, 'hr/create_job_discipline.html', {'form': form})


@admins
def job_disciplines(request):
    # Retrieve all job disciplines and annotate them with the vacancy count
    job_disciplines = JobDiscipline.objects.annotate(
        vacancy_count=Count('vacancy'))

    return render(request, 'hr/job_disciplines.html', {'job_discipline': job_disciplines})


@admins
def update_job_discipline(request, job_discipline_id):
    job_discipline = get_object_or_404(JobDiscipline, pk=job_discipline_id)

    if request.method == 'POST':
        form = JobDisciplineForm(request.POST, instance=job_discipline)
        if form.is_valid():
            form.save()
            # Redirect to the job disciplines list after updating
            return redirect('hr:job_disciplines')
    else:
        form = JobDisciplineForm(instance=job_discipline)

    return render(request, 'hr/update_job_discipline.html', {'form': form, 'job_discipline': job_discipline})


@admins
def delete_job_discipline(request, job_discipline_id):
    job_discipline = get_object_or_404(JobDiscipline, pk=job_discipline_id)

    job_discipline.delete()
    return redirect('hr:job_disciplines')


@admins
def create_certifying_body(request):
    if request.method == 'POST':
        form = CertifyingBodyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Certfying body added succesfully")
            return redirect('hr:certifying_bodies')
    else:
        form = CertifyingBodyForm()

    return render(request, 'hr/create_certifying_body.html', {'form': form})


@admins
def certifying_bodies(request):
    certifying_bodies = CertifyingBody.objects.all()
    return render(request, 'hr/certifying_bodies.html', {'c_bodies': certifying_bodies})


@admins
def edit_certifying_body(request, certifying_body_id):
    certifying_body = get_object_or_404(CertifyingBody, pk=certifying_body_id)

    if request.method == 'POST':
        form = CertifyingBodyForm(request.POST, instance=certifying_body)
        if form.is_valid():
            form.save()
            # Redirect to the certifying bodies list after updating
            return redirect('hr:certifying_bodies')
    else:
        form = CertifyingBodyForm(instance=certifying_body)

    return render(request, 'hr/edit_certifying_body.html', {'form': form, 'c_body': certifying_body})


@admins
def delete_certifying_body(request, certifying_body_id):
    certifying_body = get_object_or_404(CertifyingBody, pk=certifying_body_id)

    certifying_body.delete()
    # Redirect to the certifying bodies list after deletion
    return redirect('hr:certifying_bodies')


@admins
def create_certificate(request):
    if request.method == 'POST':
        form = CertificateForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to a success page or the certificate list page
            return redirect('hr:certificates')
    else:
        form = CertificateForm()

    return render(request, 'hr/create_certificate.html', {'form': form})


@admins
def certificates(request):
    certificates = Certificate.objects.all()
    return render(request, 'hr/certificates.html', {'certificate': certificates})


@admins
def edit_certificate(request, certificate_id):
    # Get the certificate instance to be edited
    certificate = get_object_or_404(Certificate, pk=certificate_id)

    if request.method == 'POST':
        form = CertificateForm(request.POST, instance=certificate)
        if form.is_valid():
            form.save()
            # Redirect to the certificates list page
            return redirect('hr:certificates')
    else:
        form = CertificateForm(instance=certificate)

    return render(request, 'hr/edit_certificate.html', {'form': form, 'certificate': certificate})


@admins
def delete_certificate(request, certificate_id):
    # Get the certificate instance to be deleted
    certificate = get_object_or_404(Certificate, pk=certificate_id)

    certificate.delete()
    # Redirect to the certificates list page
    return redirect('hr:certificates')


@admins
def create_field_of_study(request):
    if request.method == 'POST':
        form = FieldOfStudyForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to the list view after creating
            return redirect('hr:fields_of_study')
    else:
        form = FieldOfStudyForm()

    return render(request, 'hr/create_field_of_study.html', {'form': form})


@admins
def edit_field_of_study(request, field_of_study_id):
    field_of_study = FieldOfStudy.objects.get(pk=field_of_study_id)

    if request.method == 'POST':
        form = FieldOfStudyForm(request.POST, instance=field_of_study)
        if form.is_valid():
            form.save()
            # Redirect to the list view after editing
            return redirect('hr:fields_of_study')
    else:
        form = FieldOfStudyForm(instance=field_of_study)

    return render(request, 'hr/edit_field_of_study.html', {'form': form})


@admins
def delete_field_of_study(request, field_of_study_id):
    field_of_study = FieldOfStudy.objects.get(pk=field_of_study_id)

    field_of_study.delete()
    # Redirect to the list view after deleting
    return redirect('hr:fields_of_study')


@admins
def fields_of_study(request):
    fields_of_study = FieldOfStudy.objects.all()
    return render(request, 'hr/fields_of_study.html', {'fields_of_study': fields_of_study})


@admins
def edu_levels(request):
    edu_levels = EducationalLevel.objects.all()
    return render(request, 'hr/edu_levels.html', {'edu_levels': edu_levels})


@admins
def create_ethnicity(request):
    if request.method == 'POST':
        form = EthnicityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hr:ethnicities')
    else:
        form = EthnicityForm()
    return render(request, 'hr/create_ethnicity.html', {'form': form})


@admins
def ethnicities(request):
    ethnicities = Ethnicity.objects.all()
    return render(request, 'hr/ethnicities.html', {'ethnicities': ethnicities})


@admins
def edit_ethnicity(request, ethnicity_id):
    ethnicity = get_object_or_404(Ethnicity, id=ethnicity_id)
    if request.method == 'POST':
        form = EthnicityForm(request.POST, instance=ethnicity)
        if form.is_valid():
            form.save()
            return redirect('hr:ethnicities')
    else:
        form = EthnicityForm(instance=ethnicity)
    return render(request, 'hr/edit_ethnicity.html', {'form': form, 'ethnicity': ethnicity})


@admins
def delete_ethnicity(request, ethnicity_id):
    ethnicity = get_object_or_404(Ethnicity, id=ethnicity_id)

    ethnicity.delete()
    return redirect('hr:ethnicities')


@admins
def user_access_logs(request):
    user_logs = UserAccessLog.objects.all()
    return render(request, 'hr/user_logs.html', {'user_logs': user_logs})


@admins
def admin_access_logs(request):
    admin_logs = AdminAccessLog.objects.all()
    return render(request, 'hr/admin_logs.html', {'admin_logs': admin_logs})


@admins
def portal_reports(request):
    # Query to count users with access level 0
    access_level_0_count = CustomUser.objects.filter(access_level=0).count()

    # Query to count users with access level 1-4
    access_level_1_to_4_count = CustomUser.objects.filter(
        access_level__range=(1, 4)).count()

    # Query to count the number of vacancies created
    vacancy_count = Vacancy.objects.count()

    # Query to count the total number of applications
    total_applications_count = Application.objects.count()

    context = {
        'r_count': access_level_0_count,
        'a_count': access_level_1_to_4_count,
        'vacancy_count': vacancy_count,
        'applications_count': total_applications_count,
    }

    print(context)

    return render(request, 'hr/portal_reports.html', context)


@admins
def vacancy_report(request):
    # Calculate the count of open vacancies
    open_vacancy_count = Vacancy.objects.filter(
        Q(date_open__lte=timezone.now()) &
        Q(date_close__gt=timezone.now())
    ).count()

    # Calculate the count of closed vacancies
    closed_vacancy_count = Vacancy.objects.filter(
        Q(date_close__lte=timezone.now())
    ).count()

    # Use aggregation and annotation to count the vacancies for each job discipline
    job_disciplines_counts = JobDiscipline.objects.annotate(
        vacancy_count=Count('vacancy'))

    # Prepare the data as a list of dictionaries
    data = [
        {
            'name': discipline.name,
            'vacancy_count': discipline.vacancy_count,
        }
        for discipline in job_disciplines_counts
    ]

    # Convert data to a JSON string
    data_json = json.dumps(data)

    # Prepare data for the pie chart
    pie_chart_data = {
        'open_vacancy_count': open_vacancy_count,
        'closed_vacancy_count': closed_vacancy_count,
    }

    # Convert pie chart data to JSON
    pie_chart_data_json = json.dumps(pie_chart_data)

    all_vacancies = Vacancy.objects.all()

    return render(request, 'hr/v_report.html', {
        'job_disciplines_with_counts_json': data_json,
        'pie_chart_data_json': pie_chart_data_json,
        'jobs': all_vacancies
    })


@admins
def applications_reports(request):
    # Retrieve counts
    total_applications = Application.objects.count()
    qualify_true_count = Application.objects.filter(qualify=True).count()
    shortlisted_true_count = Application.objects.filter(
        shortlisted=True).count()

    # Retrieve counts for each vacancy
    vacancies = Vacancy.objects.all()
    vacancy_counts = []

    for vacancy in vacancies:
        vacancy_count = Application.objects.filter(vacancy=vacancy).count()
        vacancy_counts.append({
            'vacancy_title': vacancy.title,
            'application_count': vacancy_count,
        })

    # Prepare data as a dictionary
    data = {
        'labels': ['Total Applications', 'Qualify True', 'Shortlisted True'],
        'data': [total_applications, qualify_true_count, shortlisted_true_count],
        'vacancy_data': vacancy_counts,
    }

    # Convert data to JSON
    data_json = json.dumps(data)

    context = {
        'data_json': data_json,  # Add the JSON data to the context
    }

    return render(request, 'hr/application_reports.html', context)


@admins
def application_report(request, vacancy_id):
    # Retrieve the specific vacancy or return a 404 error if not found
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)

    # Retrieve counts for this specific vacancy
    total_applications = Application.objects.filter(vacancy=vacancy).count()
    qualify_true_count = Application.objects.filter(
        vacancy=vacancy, qualify=True).count()
    shortlisted_true_count = Application.objects.filter(
        vacancy=vacancy, shortlisted=True).count()

    # Prepare data for the graph as a dictionary
    graph_data = {
        'labels': ['Total Applications', 'Qualified Applicants', 'Shortlisted Applicants'],
        'data': [total_applications, qualify_true_count, shortlisted_true_count],
    }

    # Convert graph data to JSON
    graph_data_json = json.dumps(graph_data)

    context = {
        'vacancy': vacancy,
        # Add the JSON data for the graph to the context
        'graph_data_json': graph_data_json,
    }

    return render(request, 'hr/application_report.html', context)


@system_admin_hr_post_required
def adms(request):
    search_query = request.GET.get('q')
    users_with_access_5 = []

    if search_query:
        # Filter users based on search query and access level
        users_with_access_5 = CustomUser.objects.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(resume__full_name__icontains=search_query),
            access_level=5
        )

    context = {
        'search_query': search_query,
        'users_with_access_5': users_with_access_5,
    }

    return render(request, 'hr/adms.html', context)


@system_admin_required
def admin_role(request, admin_id):
    user_to_update = get_object_or_404(CustomUser, id=admin_id)

    if request.method == 'POST':
        form = UpdateFunctionForm(request.POST, instance=user_to_update)
        if form.is_valid():
            form.save()
            # Redirect to a success page or any other appropriate action
            return redirect('hr:hr_admins')  # Replace with your desired URL
    else:
        form = UpdateFunctionForm(instance=user_to_update)

    context = {
        'user_to_update': user_to_update,
        'form': form,
    }

    return render(request, 'hr/admin_role.html', context)


@system_admin_required
def hr_admin(request):
    # Filter users with access level 5 and function values 1, 2, 3, 4, or 5
    users_with_access_and_function = CustomUser.objects.filter(
        access_level=5,
        function__in=[1, 2, 3, 4, 5]
    )

    context = {
        'users_with_access_and_function': users_with_access_and_function,
    }

    return render(request, 'hr/hr.html', context)


@system_admin_required
def admin_register(request):
    if request.method == 'POST':
        form = AdminForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Construct the login link
            current_site = get_current_site(request)
            # Assuming 'login' is the name of your login URL pattern
            login_link = reverse('users:login')

            # Compose the email message
            subject = 'Account Created Successfully'
            message = f"Hello {user.username},\n\nYour admin account has been created successfully. Here are your login details:\n\nUsername: {user.username}\nEmail: {user.email}\nPassword: {form.cleaned_data['password1']}\n\nYou can now use these details to log in to your account.\n\nLogin here: {current_site}{login_link}"

            from_email = 'nelson@kenyaweb.co.ke'  # Replace with your sender email
            to_email = user.email

            # Send the email
            send_mail(subject, message, from_email, [to_email])

            # Add a success message
            messages.success(
                request, f"{user.username}'s admin account has been created successfully, and login details have been sent to {user.email}.")

            # You can perform any additional actions here after user creation, if needed
            # Redirect to a success page or another URL
            return redirect('hr:adms')
        else:
            # Form has errors, display error messages
            messages.error(
                request, "There were errors in the form. Please correct them.")
    else:
        form = AdminForm()

    return render(request, 'hr/admin_register.html', {'form': form})


@admins
def create_terms(request):
    if request.method == 'POST':
        form = TermsForm(request.POST)
        if form.is_valid():
            text = form.cleaned_data['text']
            terms, created = Terms.objects.get_or_create(
                defaults={'text': text})
            if not created:
                terms.text = text
                terms.save()
            # Redirect to the terms page or any other page
            return redirect('main:terms')
    else:
        initial_text = Terms.objects.first()
        form = TermsForm(
            initial={'text': initial_text.text if initial_text else ''})

    return render(request, 'hr/create_terms.html', {'form': form})


@system_admin_required
def import_excel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith('.xlsx'):
                return render(request, 'error_page.html', {'message': 'File is not in .xlsx format.'})

            try:
                wb = load_workbook(excel_file)
                sheet_name = 'staff list 23.10.2023'
                sheet = wb[sheet_name]

                for row in sheet.iter_rows():
                    staff_no, name, email = [cell.value for cell in row]

                    try:
                        hashed_password = make_password(
                            staff_no)  # Hash the password

                        with transaction.atomic():
                            user = CustomUser.objects.create(
                                username=staff_no,
                                email=email,
                                access_level=5,
                                password=hashed_password  # Assign the hashed password
                            )

                            # Create or update the user's resume
                            resume, _ = Resume.objects.get_or_create(user=user)
                            resume.full_name = name
                            resume.email_address = email
                            resume.save()
                    except ValidationError:
                        return render(request, 'error_page.html', {'message': 'Error during import'})

                wb.close()  # Close the workbook

                return redirect('hr:staffs')
            except (KeyError, ValidationError):
                return render(request, 'error_page.html', {'message': 'Error during import'})

    else:
        form = ExcelImportForm()

    return render(request, 'hr/import_staffs.html', {'form': form})


@system_admin_hr_post_required
def staffs(request):
    # Retrieve all staff members with access level 5
    staff_members = CustomUser.objects.filter(access_level=5)

    # Define the number of items per page
    items_per_page = 100

    # Filter staff members based on search criteria
    search_query = request.GET.get('search_query', '')
    if search_query:
        staff_members = staff_members.filter(
            Q(username__icontains=search_query) | Q(email__icontains=search_query))

    paginator = Paginator(staff_members, items_per_page)
    # Get the current page number from the request
    page = request.GET.get('page')

    staff_members = paginator.get_page(page)

    context = {
        'staff_members': staff_members,
        'search_query': search_query,
    }

    return render(request, 'hr/staffs.html', context)


@system_admin_required
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)

    if user.access_level != 5:
        return HttpResponseForbidden("Access denied")

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            # Redirect to the user list page after editing
            return redirect('hr:kgn_staffs')
    else:
        form = UserEditForm(instance=user)

    context = {
        'form': form,
        'user': user,
    }

    return render(request, 'hr/edit_staff.html', context)


@system_admin_required
def delete_staff(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)

    # Check if the user has access level 5
    if user.access_level != 5:
        return HttpResponseForbidden("Access denied")

    user.delete()
    # Redirect to the user list page after deleting
    return redirect('hr:kgn_staffs')


@system_admin_required
def reset_trials(request, user_id):
    # Get the admin user (you can implement admin authentication)
    # Get the user to reset trials for (or use get_object_or_404 to handle non-existent user)
    user_to_reset = get_object_or_404(CustomUser, id=user_id)

    # Reset the trials for the selected user
    user_to_reset.trials = 4  # Reset the trials to 3 or your desired value
    user_to_reset.is_restricted = False  # Unrestrict the user
    user_to_reset.save()

    # Add a success message
    messages.success(
        request, f"Trials reset successfully for {user_to_reset.username}. They can submit their staff No again.")

    # Redirect back to the admin panel or another appropriate page
    return redirect('hr:system_users')


@system_admin_required
def delete_users_with_access_level_5(request):
    if request.method == 'POST':
        # Delete users with access level 5
        CustomUser.objects.filter(access_level=5).delete()
        return redirect('hr:kgn_staffs')  # Redirect to a success page

    return render(request, 'hr/staffs.html')
